##excel的写入
from Z_API.Z_API_3 import http_requests
import openpyxl
class Case:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None
        self.sql=None
class Do_Excel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.workbook=openpyxl.load_workbook(file_name)
        self.sheet=self.workbook[sheet_name]
    def get_cases(self):
        max_row=self.sheet.max_row
        cases=[]
        for row in range(2,max_row+1):
            case=Case()
            case.case_id=self.sheet.cell(row=row,column=1).value
            case.title = self.sheet.cell(row=row, column=2).value
            case.url = self.sheet.cell(row=row, column=3).value
            case.data = self.sheet.cell(row=row, column=4).value
            case.method= self.sheet.cell(row=row, column=5).value
            case.expected = self.sheet.cell(row=row, column=6).value
            case.sql=self.sheet.cell(row=row, column=9).value###sql
            cases.append(case)
        self.workbook.close()
        return cases
    def Write_excel(self,row,actual,result):
        sheet=self.workbook[self.sheet_name]##打开表格
        sheet.cell(row,7).value=actual#actual值写入
        sheet.cell(row,8).value=result#result值写入
        self.workbook.save(filename=self.file_name)##写入以后要保存
        self.workbook.close()##保存以后要关闭

if __name__ == '__main__':
    do_excel=Do_Excel("Testcase.xlsx",sheet_name="login")
    case_read=do_excel.get_cases()
    http_request=http_requests.HTTPRequest()
    # print(case_read.__dict__)
    for case in case_read:
        # print(case.case_id)
        print(case.__dict__)
        ##写入的话，要先调用httprequest执行，执行成功写入
        resp=http_request.request(case.method,case.url,case.data)
        actual=resp.text
        if case.expected==actual:##判断执行厚的实际结果是否与预期结果一致
            do_excel.Write_excel(case.case_id + 1,actual,'PASS')
        else:
            do_excel.Write_excel(case.case_id + 1,actual,'FAIL')

