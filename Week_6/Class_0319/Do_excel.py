# 安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据
#
# 1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，
# 所有子列表数据放到一个大列表里面,要求把读取到的数据返回
#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值

from openpyxl import load_workbook
class do_excel:
    def read_data(self,filename,sheetname):
        #filename::目标工作部
        #sheetname:指定的表单名（操作的那个表单名）
        wb=load_workbook(filename)##打开excel
        sheet=wb[sheetname]##打开表单

        sub_list=[]##子列表
        all_data=[]
        #读取第一行
        for i in range(1,sheet.max_row+1):
            sub_list.append(sheet.cell(i,1).value)
            sub_list.append(sheet.cell(i,2).value)
            sub_list.append(sheet.cell(i,3).value)
            sub_list.append(sheet.cell(i,4).value)
            all_data.append(sub_list)
        wb.close()
        return all_data
    def write_back(self,row,column,new_value,filename,sheetname):
        wb=load_workbook(filename)
        sheet=wb[sheetname]
        #写入指定的行列值
        sheet.cell(row,column).value=new_value
        wb.save(filename)
        wb.close()

if __name__ == '__main__':
    doexcel=do_excel().read_data("py15_1bb","Sheet")
    print(doexcel)
    do_excel.write_back("py15_1bb","1","2","sddsadas")
##为啥要一行存在一起？？
#用例是一行一行的，我们用的是excel来存储测试数据

