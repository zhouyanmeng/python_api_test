# -*- coding: utf-8 -*-
# @Time    : 2018/12/29 20:51
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py

from openpyxl import load_workbook

class Cases:
    '''专门存储我们的测试数据'''
    def __init__(self):
        self.case_id=None
        self.title=None
        self.a=None
        self.b=None
        self.expected=None


class DoExcel:
    '''来获取Excel里面对应表单的数据的一个类'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#工作簿
        self.sheet_name=sheet_name#表单名
        # self.button=button

    def get_data(self):#获取数据
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        # if self.button=='all':
        #     pass
        # else:
        #     #如果是列表，比如说[1,2,3]---》 2 3 4行
        #    for i  in [1,2,3]:


        # test_data=[]#所有的数据都存在这个大列表里面  适用于方法一 方法二
        #读取数据的具体操作
        #方法一：每一行数据存在一个子列表里面
        # for i in range(2,sheet.max_row+1):#行的范围从第二行开始
        #     row_data=[]#每一行数据存在这个子列表里面
        #     row_data.append(sheet.cell(row=i,column=1).value)#存的是case_id
        #     row_data.append(sheet.cell(row=i,column=2).value)#存的是title
        #     row_data.append(sheet.cell(row=i,column=3).value)#存的是a
        #     row_data.append(sheet.cell(row=i,column=4).value)#存的是b
        #     row_data.append(sheet.cell(row=i,column=5).value)#存的是expected
        #     test_data.append(row_data)#读取完毕之后 把每一行的值存到这个test_data大列表里面去
        # return test_data

        #方法二：每一行数据存在一个子字典里面
        # for i in range(2,sheet.max_row+1):#行的范围从第二行开始
        #     row_data={}#每一行数据存在这个子字典里面
        #     row_data['case_id']=sheet.cell(row=i,column=1).value#存的是case_id
        #     row_data['title']=sheet.cell(row=i,column=2).value#存的是title
        #     row_data['a']=sheet.cell(row=i,column=3).value#存的是a
        #     row_data['b']=sheet.cell(row=i,column=4).value#存的是b
        #     row_data['expected']=sheet.cell(row=i,column=5).value#存的是expected
        #     test_data.append(row_data)#读取完毕之后 把每一行的值存到这个test_data大列表里面去
        # return test_data

        #方法三：用类与对象的思想去做---建议用这种
        case=[]#存储数据
        for i in range(2,sheet.max_row+1):#行的范围从第二行开始
            row_case=Cases()#每一行数据存在这个对象里面 Cases() 对象/实例
            row_case.case_id=sheet.cell(row=i,column=1).value#存的是case_id
            row_case.title=sheet.cell(row=i,column=2).value#存的是title
            row_case.a=sheet.cell(row=i,column=3).value#存的是a
            row_case.b=sheet.cell(row=i,column=4).value#存的是b
            row_case.expected=sheet.cell(row=i,column=5).value#存的是expected
            case.append(row_case)
        return case

    def write_back(self,row,col,value):#写回数据
        #row 行 col 列表 value你要写回的值
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        sheet.cell(row,col).value=value#写入值到单元格里面去

        wb.save(self.file_name)#要记得保存，同时要记得Excel要关闭状态



if __name__ == '__main__':
    test_data=DoExcel('python13.xlsx','add').get_data()
    print(test_data)