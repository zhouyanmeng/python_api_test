# 安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据
#
# 1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，
# 所有子列表数据放到一个大列表里面,要求把读取到的数据返回
#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值

from openpyxl import workbook
from openpyxl import load_workbook
# wb=workbook.Workbook()##创建对象
# wb.create_sheet('excel')##创建表单
# wb.save('py_HomeWork.xlsx')##保存

wb=load_workbook('py_HomeWork.xlsx')##打开excel
sheet=wb['excel']##读取表单
for i in range(1,5):
    print(sheet.cell(1,i).value)

sheet.cell(2,1,'鹅鹅鹅')##写
sheet.cell(2,2,'曲项向天歌')##写
sheet.cell(2,3,'白猫副旅长')##写
sheet.cell(2,4,'红掌拨清波')##写
wb.save('py_HomeWork.xlsx')
wb.close()





