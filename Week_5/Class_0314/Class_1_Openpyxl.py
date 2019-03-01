#####openpyel模块学习
####数据类型需要注意

#利用python去操作excel的话，就要安装pip install openpyxl
##创建excel文件的模块workbook
##读写excel文件的模块load_workbook

####操作：读写
####创建一个excel操作：：代码创建


#1手动创建：：xlsx的工作表，不支持xls结尾的工作表格
#2代码创建：：
from openpyxl import workbook ###Workbook是一个类  workbook是一个模块
#新建  workbook  sheet  cell
# wb=workbook.Workbook()#创建一个对象
# wb.create_sheet('lbb')##创建表单的方法
# wb.save('py15_1bb.xlsx')##保存，保存的文件名 def savw(self,filename)
#     ##相对地址
# ##只创建excel，不会创建自己的表单
wb=workbook.Workbook()#创建一个对象
wb.save('py15.xlsx')##保存，保存的文件名 def savw(self,filename)

#3读写操作excel
#开始操作excel
from openpyxl import load_workbook
##3步骤

wb=load_workbook('py15_1bb.xlsx')#（1）打开excel
#sheet=wb.get_sheet_by_name('sheet1')#（2）获取表单
sheet=wb['Sheet']#（2）获取表单
res=sheet.cell(1,2).value#(3)定位cell单元格，获取内容  根据行列坐标来定位单元格获取值
                            ##行列坐标都必须是数字：A--1,B--2,C--3,D--4,E--5
# print(res)
# print('res的值是：，res的数据类型是：{}'.format(res),format(type(res)))
# A1=sheet.cell(1,1).value
# print('A1的值是：，A1的数据类型是：{}'.format(A1),format(type(A1)))
# B2=sheet.cell(2,2).value
# print('B2的值是：，B2的数据类型是：{}'.format(B2),format(type(B2)))
# C3=sheet.cell(3,3).value
# print('C3的值是：，C3的数据类型是：{}'.format(C3),format(type(C3)))
# D4=sheet.cell(4,4).value
# print('D4的值是：，D4的数据类型是：{}'.format(D4),format(type(D4)))

###得出一个结果：：数字还是数字，其他数据类型都是字符串类型

#eval() 可以把数据转成python原本可以识别的数据类型
#列表，字典，元组都是字符串，用eval(),拿到的值就是list,dict,# tuple

#4写入值，保存
#(1)打开
#(2)定位单元格
#(3)写入值   保存，另存为   如果保存到当前的excel的话，记得要关闭Excel,不然汇报permission definded
wb=load_workbook('py15_1bb.xlsx')#（1）打开excel
sheet=wb['Sheet']#（2）获取表单
sheet.cell(5,1).value='白日依山尽'##赋值运算   3赋值
sheet.cell(5,2,"黄河入海流")##(单元格，直接写入)
wb.save('py16_1bb.xlsx')###不存在的就新建
wb.close()###手动打开，手动关闭，，，操作完毕完成，一定要关闭文件
            ###如果保存到当前的excel的话，记得要关闭Excel,不然汇报permission definded
#5循环读值


