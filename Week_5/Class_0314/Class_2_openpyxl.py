from openpyxl import load_workbook
wb=load_workbook('py15_1bb.xlsx')
sheet=wb['sheet']
print(sheet.cell(2,1).value)
print(sheet.cell(2,2).value)
print(sheet.cell(2,3).value)
print(sheet.cell(2,4).value)

for i in range(1,5):
    print(sheet.cell(2, i).value)


for i in range(1,sheet.max_column+1):###取头不取尾
    print(sheet.cell(2, i).value)
###读取某一行的值
print(sheet.max_row)
print(sheet.max_column)

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        if sheet.cell(i,j).value:###非空为真，才执行下面的代码
            print(sheet.cell(i, i).value)

##单元格为空的时候，读取的值是none