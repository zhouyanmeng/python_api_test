# 安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据
#
# 1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，
# 所有子列表数据放到一个大列表里面,要求把读取到的数据返回
#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值
##为啥要一行存在一起？？
#用例是一行一行的，我们用的是excel来存储测试数据
from openpyxl import load_workbook
class do_excel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
    def read_data(self):
        #filename::目标工作部
        #sheetname:指定的表单名（操作的那个表单名）
        wb=load_workbook(self.filename)##打开excel
        sheet=wb[self.sheetname]##打开表单

        sub_list=[]##子列表
        all_data=[]
        #读取第一行
        for i in range(1,sheet.max_row+1):
            sub_list.append(sheet.cell(i,1).value)
            sub_list.append(sheet.cell(i,2).value)
            sub_list.append(sheet.cell(i,3).value)
            sub_list.append(sheet.cell(i,4).value)
            all_data.append(sub_list)
        wb.close()
        return all_data
    def write_back(self,row,column,new_value):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheetname]
        #写入指定的行列值
        sheet.cell(row,column).value=new_value
        wb.save(self.filename)
        wb.close()

if __name__ == '__main__':
    # read=do_excel("py15_1bb.xlsx","Sheet")
    # read_data_new1=read.read_data()
    # print(read_data_new1)

    t=do_excel("py15_1bb.xlsx","Sheet")
    t.write_back(1,2,"周五来上课")
    all_data=t.read_data()
    print(all_data)