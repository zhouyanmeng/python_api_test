##excel的数据读取
# 类的对象实例化去获取sheet里面单而每一个cell的值
import openpyxl
class Case:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None
class Do_Excel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.workbook=openpyxl.load_workbook(file_name)
        self.sheet=self.workbook[sheet_name]
    def get_cases(self):
        max_row=self.sheet.max_row
        cases=[]
        for row in range(2,max_row+1):
            case=Case()
            case.case_id=self.sheet.cell(row=row,column=1).value
            case.title = self.sheet.cell(row=row, column=2).value
            case.url = self.sheet.cell(row=row, column=3).value
            case.data = self.sheet.cell(row=row, column=4).value
            case.method= self.sheet.cell(row=row, column=5).value
            case.expected = self.sheet.cell(row=row, column=6).value
            cases.append(case)
        return cases
if __name__ == '__main__':
    do_ecel=Do_Excel("Testcase.xlsx",sheet_name="login")
    case_read=do_ecel.get_cases()
    # print(case_read.__dict__)
    for case in case_read:
        # print(case.case_id)
        # print(case.title)
        # print(case.url)
        # print(case.data)
        # print(case.method)
        # print(case.expected)
        print(case.__dict__)
