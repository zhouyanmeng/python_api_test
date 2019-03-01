##通过行列去获取sheet里面每一个cell的值
import openpyxl
class Do_Excel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.workbook=openpyxl.load_workbook(file_name)
        self.sheet=self.workbook[sheet_name]
    def get_case(self):
        max_row=self.sheet.max_row
        cases=[]
        for row in range(2,max_row+1):
            case={}
            case["case_id"]=self.sheet.cell(row=row,column=1).value
            case["title"] = self.sheet.cell(row=row, column=2).value
            case["url"] = self.sheet.cell(row=row, column=3).value
            case["data"] = self.sheet.cell(row=row, column=4).value
            case["method"] = self.sheet.cell(row=row, column=5).value
            case["excepted"] = self.sheet.cell(row=row, column=6).value
            cases.append(case)
        return cases
if __name__ == '__main__':
    case_read=Do_Excel("Testcase.xlsx","login").get_case()
    print(case_read)
